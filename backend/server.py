from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime
import base64
import shutil
from openpyxl import load_workbook
import aiosmtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import json

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'qse_database')]

# Create the main app
app = FastAPI(title="QSE Industrial App API")

# Create router with /api prefix
api_router = APIRouter(prefix="/api")

# Excel template path
EXCEL_TEMPLATE_PATH = ROOT_DIR.parent / "excel_template.xlsm"
GENERATED_FILES_PATH = ROOT_DIR / "generated_files"
GENERATED_FILES_PATH.mkdir(exist_ok=True)

# ===================== MODELS =====================

class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str  # Employee code
    name: str
    first_name: str
    service: str
    is_admin: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)

class UserCreate(BaseModel):
    code: str
    name: str
    first_name: str
    service: str
    is_admin: bool = False

class UserLogin(BaseModel):
    code: str

class Photo(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    data: str  # Base64 encoded
    filename: str
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ActionCorrective(BaseModel):
    action: str
    responsable: str
    delai: Optional[str] = None
    type_action: Optional[str] = None
    statut: str = "A lancer"

class FicheQSE(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: str  # "Qualité", "Sécurité", "Environnement"
    
    # Common fields
    date_evenement: datetime
    heure_evenement: str
    constate_par: str
    service_emetteur: str
    service_concerne: Optional[str] = None
    
    # Quality specific
    non_conformite_constatee: Optional[str] = None
    defaut: Optional[str] = None
    ccp_prpo: Optional[str] = None
    categorie_corps_etranger: Optional[str] = None
    quantite_concernee: Optional[str] = None
    
    # Product traceability
    produit: Optional[str] = None
    grammage: Optional[str] = None
    marque: Optional[str] = None
    ligne: Optional[str] = None
    ddm: Optional[str] = None
    quantieme: Optional[str] = None
    heure_production: Optional[str] = None
    
    # Other traceability
    numero_lot: Optional[str] = None
    numero_palette: Optional[str] = None
    code_sca: Optional[str] = None
    reference_interne: Optional[str] = None
    date_production: Optional[str] = None
    numero_bobine: Optional[str] = None
    autres_tracabilite: Optional[str] = None
    
    # Description
    description: str
    
    # Criticality
    criticite: str  # "Mineure", "Majeure", "Critique"
    impact_securite_aliments: Optional[str] = None
    
    # Treatment (Quality)
    traitement_blocage: bool = False
    traitement_methanisation: bool = False
    traitement_fonte: bool = False
    traitement_analyses: bool = False
    traitement_alimentation_animale: bool = False
    traitement_autres: Optional[str] = None
    date_traitement: Optional[str] = None
    nom_traitement: Optional[str] = None
    
    # Safety specific
    type_incident: Optional[str] = None  # "Presqu'accident", "Situation dangereuse", etc.
    type_risque: Optional[str] = None
    regle_or: Optional[str] = None
    
    # Environment specific
    type_env: Optional[str] = None  # "Eaux", "Air", "Sol", "Déchets", etc.
    traitement_env: Optional[List[str]] = None
    
    # Causes (5M)
    cause_main_oeuvre: Optional[str] = None
    cause_materiel: Optional[str] = None
    cause_methode: Optional[str] = None
    cause_milieu: Optional[str] = None
    cause_matiere: Optional[str] = None
    
    # Corrective actions
    actions_correctives: List[ActionCorrective] = []
    
    # Photos
    photos: List[Photo] = []
    
    # Signature
    signature: Optional[str] = None  # Base64 encoded
    
    # Status
    statut: str = "Brouillon"  # "Brouillon", "Validé", "Envoyé", "Échec d'envoi"
    
    # Metadata
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: str
    excel_filename: Optional[str] = None
    sync_status: str = "pending"  # "pending", "synced", "failed"

class FicheCreate(BaseModel):
    type: str
    date_evenement: datetime
    heure_evenement: str
    constate_par: str
    service_emetteur: str
    service_concerne: Optional[str] = None
    non_conformite_constatee: Optional[str] = None
    defaut: Optional[str] = None
    ccp_prpo: Optional[str] = None
    categorie_corps_etranger: Optional[str] = None
    quantite_concernee: Optional[str] = None
    produit: Optional[str] = None
    grammage: Optional[str] = None
    marque: Optional[str] = None
    ligne: Optional[str] = None
    ddm: Optional[str] = None
    quantieme: Optional[str] = None
    heure_production: Optional[str] = None
    numero_lot: Optional[str] = None
    numero_palette: Optional[str] = None
    code_sca: Optional[str] = None
    reference_interne: Optional[str] = None
    date_production: Optional[str] = None
    numero_bobine: Optional[str] = None
    autres_tracabilite: Optional[str] = None
    description: str
    criticite: str
    impact_securite_aliments: Optional[str] = None
    traitement_blocage: bool = False
    traitement_methanisation: bool = False
    traitement_fonte: bool = False
    traitement_analyses: bool = False
    traitement_alimentation_animale: bool = False
    traitement_autres: Optional[str] = None
    date_traitement: Optional[str] = None
    nom_traitement: Optional[str] = None
    type_incident: Optional[str] = None
    type_risque: Optional[str] = None
    regle_or: Optional[str] = None
    type_env: Optional[str] = None
    traitement_env: Optional[List[str]] = None
    cause_main_oeuvre: Optional[str] = None
    cause_materiel: Optional[str] = None
    cause_methode: Optional[str] = None
    cause_milieu: Optional[str] = None
    cause_matiere: Optional[str] = None
    actions_correctives: List[ActionCorrective] = []
    photos: List[Photo] = []
    signature: Optional[str] = None
    created_by: str

class EmailConfig(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    smtp_server: str = "smtp.gmail.com"
    smtp_port: int = 587
    smtp_user: str = ""
    smtp_password: str = ""
    use_tls: bool = True
    default_recipients: List[str] = []
    auto_recipients_by_service: Dict[str, List[str]] = {}

class ConfigData(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    services: List[str] = []
    non_conformites: List[str] = []
    defauts: List[str] = []
    ccp_prpo: List[str] = []
    categories_corps_etranger: List[str] = []
    types_risque: List[str] = []
    regles_or: List[str] = []
    types_env: List[str] = []
    lieux: List[str] = []
    postes: List[str] = []
    types_action: List[str] = []
    statuts_action: List[str] = []

# ===================== DEFAULT DATA =====================

DEFAULT_CONFIG = {
    "services": [
        "REPC", "Fabrication PPC", "Affinage PPC", "Conditionnement PPC",
        "Expéditions PPC", "Fabrication PPNC", "Affinage PPNC", "Conditionnement PPNC",
        "Expéditions PPNC", "Maintenance générale", "Maintenance Fab/Aff PPC",
        "Maintenance Condi PPC", "Maintenance Fab PPNC", "Maintenance Condi PPNC",
        "Maintenance REPC", "Magasin", "Laboratoire", "Collecte", "Administratif",
        "Garage", "Rétrocession", "Froid ferme", "Qualité", "Environnement", "Sécurité", "Autre"
    ],
    "non_conformites": [
        "Situation", "Matière première / Ingrédient", "Coproduit",
        "Produit en cours", "Produit Fini", "Emballage", "Echantillon"
    ],
    "defauts": [
        "Qualité - Autre", "Moisissures", "Goût - texture - odeur", "Aspect",
        "Corps étranger", "Microbiologie", "Physico-chimie", "Emballage - marquage",
        "Poids", "Nuisibles"
    ],
    "ccp_prpo": [
        "CCP - Antibiotiques", "CCP - Pasteurisation", "PRPo - Préparation milieu ferment",
        "CCP - DPM", "PRPo - Etanchéité", "PRPo - Composition gazeuse"
    ],
    "categories_corps_etranger": [
        "Plastique dur", "Plastique divers", "Papier/Carton", "Verre",
        "Graisse", "Métal", "Encre/marquage", "Nuisible", "Fromage",
        "Bois", "Cheveu", "Autre"
    ],
    "types_risque": [
        "Agents chimiques dangereux", "Ambiances climatiques / températures extrêmes",
        "ATEX - risque explosion", "Bruit", "Brûlure thermique",
        "Chute avec dénivellation", "Chute de hauteur", "Chute de plain-pied",
        "Chute d'objet", "Circulation", "Coincement / Ecrasement", "Coupure",
        "Ecran", "Electricité", "Espaces confinés", "Gestes répétitifs",
        "Incendie", "Machines", "Manutention manuelle", "Manutention mécanique",
        "Outils à main", "Postures pénibles", "Risque biologique", "Risque routier",
        "Risques physiques", "Risques psychosociaux", "Travail de nuit",
        "Travail en équipes alternantes", "Vibrations mécaniques", "Autre"
    ],
    "regles_or": [
        "EPI", "Machines", "Consignation", "Espace confiné / Travail en hauteur",
        "Produit chimique", "Manutention manuelle / Posture", "Conduite d'engins",
        "Circulation piétonne", "Circulation routière", "NON APPLICABLE"
    ],
    "types_env": [
        "Eaux (fuite : eaux usées, eaux pluviales, eaux de ville...)",
        "Air : fuite, rejet (chaudière, TAR)",
        "Sol (déversement au sol)",
        "Déchets (tri …)",
        "Autres (réglementaire…)"
    ],
    "lieux": [
        "cuve", "pressage", "GSV", "Acidification", "Laverie", "Saumure",
        "filtre/pasto saumure", "salle levains", "Sortie saumure", "sds", "cave",
        "Quai", "Local chargeur", "Autres", "E301", "TANK de Mat", "Chariot",
        "combles", "locaux sociaux", "Galerie", "Manip meules", "Démouleuse",
        "Frigo cdt", "Autre"
    ],
    "postes": [
        "salle de soins", "sortie saumure", "conduite moulage", "conduite cuves",
        "cariste appro sds", "cariste sortie saumure", "cariste cave", "gradeur",
        "nettoyage fab", "expedition camion", "prepa camions", "autre"
    ],
    "types_action": ["Humaine", "Organisationnelle", "Technique"],
    "statuts_action": ["Close", "En cours", "A lancer"]
}

# ===================== ROUTES =====================

@api_router.get("/")
async def root():
    return {"message": "QSE Industrial App API", "version": "1.0.0"}

# ----- Users -----

@api_router.post("/users", response_model=User)
async def create_user(user: UserCreate):
    user_dict = user.model_dump()
    user_obj = User(**user_dict)
    await db.users.insert_one(user_obj.model_dump())
    return user_obj

@api_router.post("/users/login")
async def login_user(login: UserLogin):
    user = await db.users.find_one({"code": login.code})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    return User(**user)

@api_router.get("/users", response_model=List[User])
async def get_users():
    users = await db.users.find().to_list(1000)
    return [User(**user) for user in users]

@api_router.get("/users/{user_id}", response_model=User)
async def get_user(user_id: str):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    return User(**user)

# ----- Fiches QSE -----

@api_router.post("/fiches", response_model=FicheQSE)
async def create_fiche(fiche: FicheCreate):
    fiche_dict = fiche.model_dump()
    fiche_obj = FicheQSE(**fiche_dict)
    await db.fiches.insert_one(fiche_obj.model_dump())
    return fiche_obj

@api_router.get("/fiches", response_model=List[FicheQSE])
async def get_fiches(
    statut: Optional[str] = None,
    type: Optional[str] = None,
    service: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    query = {}
    if statut:
        query["statut"] = statut
    if type:
        query["type"] = type
    if service:
        query["service_emetteur"] = service
    
    fiches = await db.fiches.find(query).sort("created_at", -1).to_list(1000)
    return [FicheQSE(**fiche) for fiche in fiches]

@api_router.get("/fiches/{fiche_id}", response_model=FicheQSE)
async def get_fiche(fiche_id: str):
    fiche = await db.fiches.find_one({"id": fiche_id})
    if not fiche:
        raise HTTPException(status_code=404, detail="Fiche non trouvée")
    return FicheQSE(**fiche)

@api_router.put("/fiches/{fiche_id}", response_model=FicheQSE)
async def update_fiche(fiche_id: str, fiche: FicheCreate):
    existing = await db.fiches.find_one({"id": fiche_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Fiche non trouvée")
    
    fiche_dict = fiche.model_dump()
    fiche_dict["id"] = fiche_id
    fiche_dict["updated_at"] = datetime.utcnow()
    fiche_dict["created_at"] = existing["created_at"]
    
    await db.fiches.update_one({"id": fiche_id}, {"$set": fiche_dict})
    updated = await db.fiches.find_one({"id": fiche_id})
    return FicheQSE(**updated)

@api_router.delete("/fiches/{fiche_id}")
async def delete_fiche(fiche_id: str):
    result = await db.fiches.delete_one({"id": fiche_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fiche non trouvée")
    return {"message": "Fiche supprimée"}

# ----- Validate and Generate Excel -----

@api_router.post("/fiches/{fiche_id}/validate")
async def validate_fiche(fiche_id: str):
    fiche = await db.fiches.find_one({"id": fiche_id})
    if not fiche:
        raise HTTPException(status_code=404, detail="Fiche non trouvée")
    
    fiche_obj = FicheQSE(**fiche)
    
    # Generate Excel file
    try:
        excel_filename = await generate_excel(fiche_obj)
        
        # Update fiche status
        await db.fiches.update_one(
            {"id": fiche_id},
            {"$set": {
                "statut": "Validé",
                "excel_filename": excel_filename,
                "updated_at": datetime.utcnow()
            }}
        )
        
        return {
            "message": "Fiche validée et Excel généré",
            "excel_filename": excel_filename,
            "statut": "Validé"
        }
    except Exception as e:
        logging.error(f"Error generating Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur génération Excel: {str(e)}")

def safe_write_cell(ws, cell_coord: str, value):
    """Safely write to a cell, handling merged cells"""
    if value is None:
        return
    
    # Find if cell is in a merged range
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            # Get top-left cell of merged range
            top_left = str(merged_range).split(':')[0]
            ws[top_left] = value
            return
    
    # Not merged, write directly
    ws[cell_coord] = value

async def generate_excel(fiche: FicheQSE) -> str:
    """Generate Excel file from fiche data"""
    
    # Check if template exists
    if not EXCEL_TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="Template Excel non trouvé")
    
    # Load template with macros
    wb = load_workbook(EXCEL_TEMPLATE_PATH, keep_vba=True)
    
    # Get the correct sheet based on type
    if fiche.type == "Qualité":
        ws = wb["ENS Qualité"]
    elif fiche.type == "Sécurité":
        ws = wb["ENS Sécurité"]
    else:
        ws = wb["ENS Environnement Energie"]
    
    # Fill common data
    date_str = fiche.date_evenement.strftime("%d/%m/%Y")
    
    if fiche.type == "Qualité":
        # Quality sheet specific cells
        safe_write_cell(ws, "E6", date_str)
        safe_write_cell(ws, "G6", fiche.heure_evenement)
        safe_write_cell(ws, "L6", fiche.constate_par)
        safe_write_cell(ws, "E7", fiche.service_emetteur)
        safe_write_cell(ws, "M7", fiche.service_concerne)
        
        # Identification
        safe_write_cell(ws, "G9", fiche.non_conformite_constatee)
        safe_write_cell(ws, "G10", fiche.defaut)
        safe_write_cell(ws, "G11", fiche.ccp_prpo)
        safe_write_cell(ws, "G12", fiche.categorie_corps_etranger)
        safe_write_cell(ws, "G13", fiche.quantite_concernee)
        
        # Traceability
        safe_write_cell(ws, "G14", fiche.produit)
        safe_write_cell(ws, "L14", fiche.numero_lot)
        safe_write_cell(ws, "G15", fiche.grammage)
        safe_write_cell(ws, "L15", fiche.numero_palette)
        safe_write_cell(ws, "G16", fiche.marque)
        safe_write_cell(ws, "L16", fiche.code_sca)
        safe_write_cell(ws, "G17", fiche.ligne)
        safe_write_cell(ws, "L17", fiche.reference_interne)
        safe_write_cell(ws, "G18", fiche.ddm)
        safe_write_cell(ws, "L18", fiche.date_production)
        safe_write_cell(ws, "G19", fiche.quantieme)
        safe_write_cell(ws, "L19", fiche.numero_bobine)
        safe_write_cell(ws, "G20", fiche.heure_production)
        safe_write_cell(ws, "L20", fiche.autres_tracabilite)
        
        # Description
        safe_write_cell(ws, "D22", fiche.description)
        
        # Criticality
        if fiche.criticite == "Mineure":
            safe_write_cell(ws, "H30", "X")
        elif fiche.criticite == "Majeure":
            safe_write_cell(ws, "L30", "X")
        elif fiche.criticite == "Critique":
            safe_write_cell(ws, "N30", "X")
        
        safe_write_cell(ws, "N31", fiche.impact_securite_aliments)
        
        # Treatment
        if fiche.traitement_blocage:
            safe_write_cell(ws, "E35", "X")
        if fiche.traitement_methanisation:
            safe_write_cell(ws, "J35", "X")
        if fiche.traitement_fonte:
            safe_write_cell(ws, "E37", "X")
        if fiche.traitement_analyses:
            ws["J37"] = "X"
        if fiche.traitement_alimentation_animale:
            ws["E39"] = "X"
        if fiche.traitement_autres:
            ws["J39"] = fiche.traitement_autres
        
        ws["E41"] = fiche.date_traitement
        ws["L41"] = fiche.nom_traitement
        
        # Causes (5M)
        ws["G43"] = fiche.cause_main_oeuvre
        ws["G44"] = fiche.cause_materiel
        ws["G45"] = fiche.cause_methode
        ws["G46"] = fiche.cause_milieu
        ws["G47"] = fiche.cause_matiere
        
        # Corrective actions
        for i, action in enumerate(fiche.actions_correctives[:5]):
            row = 51 + i
            ws[f"E{row}"] = action.action
            ws[f"J{row}"] = action.responsable
            ws[f"L{row}"] = action.delai
            ws[f"N{row}"] = action.statut
    
    elif fiche.type == "Sécurité":
        # Safety sheet
        ws["E6"] = date_str
        ws["L6"] = fiche.constate_par
        ws["E7"] = fiche.service_emetteur
        
        # Type d'incident (checkboxes)
        if fiche.type_incident == "Presqu'accident":
            ws["E9"] = "X"
        elif fiche.type_incident == "Risques psychosociaux":
            ws["N9"] = "X"
        elif fiche.type_incident == "Situation dangereuse":
            ws["E12"] = "X"
        elif fiche.type_incident == "Acte dangereux":
            ws["E14"] = "X"
        elif fiche.type_incident == "Impact environnemental":
            ws["N15"] = "X"
        
        # Description
        ws["D17"] = fiche.description
        
        # Règle d'or
        ws["G24"] = fiche.regle_or
        
        # Causes
        ws["D26"] = f"Main d'œuvre: {fiche.cause_main_oeuvre or ''}\nMatériel: {fiche.cause_materiel or ''}\nMéthode: {fiche.cause_methode or ''}\nMilieu: {fiche.cause_milieu or ''}"
        
        # Actions
        for i, action in enumerate(fiche.actions_correctives[:3]):
            if i == 0:
                ws["E31"] = action.action
            else:
                ws["E32"] = action.action
            ws[f"J{31+i}"] = action.responsable
            ws[f"L{31+i}"] = action.delai
        
        # Redaction date
        ws["E35"] = date_str
        ws["L35"] = fiche.constate_par
    
    else:  # Environment
        ws["E6"] = date_str
        ws["L6"] = fiche.constate_par
        ws["E7"] = fiche.service_emetteur
        
        # Type environnement
        if fiche.type_env and "Eaux" in fiche.type_env:
            ws["E10"] = "X"
        if fiche.type_env and "Air" in fiche.type_env:
            ws["E13"] = "X"
        if fiche.type_env and "Sol" in fiche.type_env:
            ws["E15"] = "X"
        if fiche.type_env and "Déchets" in fiche.type_env:
            ws["N10"] = "X"
        
        # Description
        ws["D18"] = fiche.description
        
        # Criticality
        if fiche.criticite == "Mineure":
            ws["H26"] = "X"
        elif fiche.criticite == "Majeure":
            ws["L26"] = "X"
        elif fiche.criticite == "Critique":
            ws["N26"] = "X"
        
        # Causes
        ws["G35"] = fiche.cause_main_oeuvre
        ws["G36"] = fiche.cause_materiel
        ws["G37"] = fiche.cause_methode
        ws["G38"] = fiche.cause_milieu
        ws["G39"] = fiche.cause_matiere
        
        # Redaction
        ws["E46"] = date_str
        ws["L46"] = fiche.constate_par
    
    # Generate filename
    service_clean = fiche.service_emetteur.replace(" ", "_").replace("/", "-")
    date_file = fiche.date_evenement.strftime("%Y%m%d")
    heure_file = fiche.heure_evenement.replace(":", "")
    filename = f"NC_{service_clean}_{date_file}_{heure_file}.xlsm"
    
    filepath = GENERATED_FILES_PATH / filename
    wb.save(filepath)
    wb.close()
    
    return filename

# ----- Download Excel -----

@api_router.get("/fiches/{fiche_id}/download")
async def download_excel(fiche_id: str):
    fiche = await db.fiches.find_one({"id": fiche_id})
    if not fiche:
        raise HTTPException(status_code=404, detail="Fiche non trouvée")
    
    if not fiche.get("excel_filename"):
        raise HTTPException(status_code=400, detail="Excel non encore généré")
    
    filepath = GENERATED_FILES_PATH / fiche["excel_filename"]
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Fichier Excel non trouvé")
    
    return FileResponse(
        path=str(filepath),
        filename=fiche["excel_filename"],
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
    )

# ----- Send Email -----

@api_router.post("/fiches/{fiche_id}/send-email")
async def send_fiche_email(fiche_id: str):
    fiche = await db.fiches.find_one({"id": fiche_id})
    if not fiche:
        raise HTTPException(status_code=404, detail="Fiche non trouvée")
    
    fiche_obj = FicheQSE(**fiche)
    
    # Get email config
    email_config = await db.email_config.find_one()
    if not email_config or not email_config.get("smtp_user"):
        # Update status to "Validé" (email not configured)
        await db.fiches.update_one(
            {"id": fiche_id},
            {"$set": {"statut": "Validé", "updated_at": datetime.utcnow()}}
        )
        return {
            "message": "Configuration email non définie. Fiche validée mais email non envoyé.",
            "statut": "Validé"
        }
    
    config = EmailConfig(**email_config)
    
    # Get recipients
    recipients = list(config.default_recipients)
    service_recipients = config.auto_recipients_by_service.get(fiche_obj.service_emetteur, [])
    recipients.extend(service_recipients)
    recipients = list(set(recipients))  # Remove duplicates
    
    if not recipients:
        await db.fiches.update_one(
            {"id": fiche_id},
            {"$set": {"statut": "Validé", "updated_at": datetime.utcnow()}}
        )
        return {"message": "Aucun destinataire configuré", "statut": "Validé"}
    
    try:
        # Create email
        msg = MIMEMultipart()
        date_str = fiche_obj.date_evenement.strftime("%d/%m/%Y")
        
        msg["Subject"] = f"Déclaration non-conformité – {fiche_obj.service_emetteur} – {date_str} {fiche_obj.heure_evenement}"
        msg["From"] = config.smtp_user
        msg["To"] = ", ".join(recipients)
        
        # Email body
        body = f"""
Bonjour,

Une nouvelle fiche de non-conformité a été déclarée :

Type : {fiche_obj.type}
Date : {date_str}
Heure : {fiche_obj.heure_evenement}
Service : {fiche_obj.service_emetteur}
Constaté par : {fiche_obj.constate_par}
Criticité : {fiche_obj.criticite}

Description :
{fiche_obj.description}

Le fichier Excel est joint à ce mail.

Cordialement,
Application QSE Mobile
        """
        msg.attach(MIMEText(body, "plain"))
        
        # Attach Excel file
        if fiche_obj.excel_filename:
            filepath = GENERATED_FILES_PATH / fiche_obj.excel_filename
            if filepath.exists():
                with open(filepath, "rb") as f:
                    part = MIMEBase("application", "vnd.ms-excel.sheet.macroEnabled.12")
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header("Content-Disposition", f"attachment; filename={fiche_obj.excel_filename}")
                    msg.attach(part)
        
        # Attach photos
        for i, photo in enumerate(fiche_obj.photos):
            photo_data = base64.b64decode(photo.data)
            part = MIMEBase("image", "jpeg")
            part.set_payload(photo_data)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename=photo_{i+1}.jpg")
            msg.attach(part)
        
        # Send email
        await aiosmtplib.send(
            msg,
            hostname=config.smtp_server,
            port=config.smtp_port,
            username=config.smtp_user,
            password=config.smtp_password,
            use_tls=config.use_tls
        )
        
        # Update status
        await db.fiches.update_one(
            {"id": fiche_id},
            {"$set": {"statut": "Envoyé", "updated_at": datetime.utcnow()}}
        )
        
        return {"message": "Email envoyé avec succès", "statut": "Envoyé"}
        
    except Exception as e:
        logging.error(f"Error sending email: {e}")
        await db.fiches.update_one(
            {"id": fiche_id},
            {"$set": {"statut": "Échec d'envoi", "updated_at": datetime.utcnow()}}
        )
        return {"message": f"Erreur envoi email: {str(e)}", "statut": "Échec d'envoi"}

# ----- Sync (for offline) -----

@api_router.post("/sync")
async def sync_fiches(fiches: List[FicheCreate]):
    """Sync multiple fiches from offline storage"""
    results = []
    for fiche in fiches:
        try:
            fiche_dict = fiche.model_dump()
            fiche_obj = FicheQSE(**fiche_dict)
            await db.fiches.insert_one(fiche_obj.model_dump())
            results.append({"id": fiche_obj.id, "status": "synced"})
        except Exception as e:
            results.append({"status": "failed", "error": str(e)})
    
    return {"results": results}

# ----- Configuration -----

@api_router.get("/config")
async def get_config():
    config = await db.config.find_one()
    if not config:
        # Return default config
        return DEFAULT_CONFIG
    return ConfigData(**config).model_dump()

@api_router.put("/config")
async def update_config(config: ConfigData):
    existing = await db.config.find_one()
    if existing:
        await db.config.update_one({"id": existing["id"]}, {"$set": config.model_dump()})
    else:
        await db.config.insert_one(config.model_dump())
    return config

@api_router.get("/email-config")
async def get_email_config():
    config = await db.email_config.find_one()
    if not config:
        return EmailConfig().model_dump()
    # Don't return password
    config["smtp_password"] = "***" if config.get("smtp_password") else ""
    return config

@api_router.put("/email-config")
async def update_email_config(config: EmailConfig):
    existing = await db.email_config.find_one()
    config_dict = config.model_dump()
    
    # If password is masked, keep the old one
    if config.smtp_password == "***" and existing:
        config_dict["smtp_password"] = existing.get("smtp_password", "")
    
    if existing:
        await db.email_config.update_one({"id": existing["id"]}, {"$set": config_dict})
    else:
        await db.email_config.insert_one(config_dict)
    
    return {"message": "Configuration email mise à jour"}

# ----- Stats -----

@api_router.get("/stats")
async def get_stats():
    total = await db.fiches.count_documents({})
    brouillon = await db.fiches.count_documents({"statut": "Brouillon"})
    valide = await db.fiches.count_documents({"statut": "Validé"})
    envoye = await db.fiches.count_documents({"statut": "Envoyé"})
    echec = await db.fiches.count_documents({"statut": "Échec d'envoi"})
    
    by_type = {
        "Qualité": await db.fiches.count_documents({"type": "Qualité"}),
        "Sécurité": await db.fiches.count_documents({"type": "Sécurité"}),
        "Environnement": await db.fiches.count_documents({"type": "Environnement"})
    }
    
    return {
        "total": total,
        "by_status": {
            "brouillon": brouillon,
            "valide": valide,
            "envoye": envoye,
            "echec": echec
        },
        "by_type": by_type
    }

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
